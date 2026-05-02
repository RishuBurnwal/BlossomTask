import json
import os
import csv
import argparse
import sys
from pathlib import Path
from runtime_config import load_root_env
from datetime import datetime

import requests

# ── Optional openpyxl for Excel output ──────────────────────────────────────
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ── Constants ────────────────────────────────────────────────────────────────
TIMEOUT_SECONDS = 120          # generous timeout for large payloads
SCRIPT_NAME     = "GetTask"
SCRIPTS_DIR     = Path(__file__).resolve().parent
OUTPUT_DIR      = SCRIPTS_DIR / "outputs" / SCRIPT_NAME

# Output file paths
CSV_PATH     = OUTPUT_DIR / "data.csv"
EXCEL_PATH   = OUTPUT_DIR / "data.xlsx"
PAYLOAD_PATH = OUTPUT_DIR / "payload.json"
LOGS_PATH    = OUTPUT_DIR / "logs.txt"
QUERY_PATH   = OUTPUT_DIR / "query.txt"
CLOSING_OUTPUT_DIR = SCRIPTS_DIR / "outputs" / "ClosingTask"
CLOSING_LOGS_PATH = CLOSING_OUTPUT_DIR / "logs.txt"
CLOSING_LOGS_BY_DATE_DIR = CLOSING_OUTPUT_DIR / "logs_by_date"


# ── Helpers ──────────────────────────────────────────────────────────────────

def load_dotenv_file(path=None):
    """Load environment variables from the root .env file."""
    load_root_env(Path(path) if path is not None else None)


def _required_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise SystemExit(f"[GetTask] Missing required environment variable: {name}")
    return value


def _normalized_url(raw_url: str) -> str:
    """Fix common URL typo: ':8061api/' → ':8061/api/'"""
    return raw_url.replace(":8061api/", ":8061/api/")


def get_now_iso() -> str:
    return datetime.now().isoformat()


def configure_console_encoding() -> None:
    """Avoid Windows cp1252 crashes when printing Unicode log characters."""
    if not sys.platform.startswith("win"):
        return

    try:
        stdout_encoding = (getattr(sys.stdout, "encoding", "") or "").lower()
        stderr_encoding = (getattr(sys.stderr, "encoding", "") or "").lower()

        if hasattr(sys.stdout, "reconfigure") and stdout_encoding and not stdout_encoding.startswith("utf"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure") and stderr_encoding and not stderr_encoding.startswith("utf"):
            sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        # Keep script running even when stream reconfiguration is unavailable.
        pass


# ── logs.txt helpers ─────────────────────────────────────────────────────────

def load_logged_ids() -> set:
    """
    Read logs.txt and return a set of already-processed order IDs.
    If the file does not exist, return an empty set (process everything).
    """
    if not LOGS_PATH.exists():
        return set()
    ids = set()
    with open(LOGS_PATH, "r", encoding="utf-8") as f:
        for line in f:
            oid = line.strip()
            if oid:
                ids.add(oid)
    return ids


def append_logged_id(order_id: str):
    """Append a single order ID to logs.txt."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(LOGS_PATH, "a", encoding="utf-8") as f:
        f.write(order_id + "\n")


def _normalize_order_id(value) -> str:
    order_id = str(value or "").strip()
    if order_id.endswith(".0") and order_id[:-2].isdigit():
        return order_id[:-2]
    return order_id


def _load_order_ids_from_log(path: Path) -> set:
    if not path.exists() or not path.is_file():
        return set()
    ids = set()
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            order_id = _normalize_order_id(line)
            if order_id:
                ids.add(order_id)
    return ids


def load_closed_order_ids() -> set:
    """Return order IDs already completed by ClosingTask."""
    closed_ids = set()
    closed_ids.update(_load_order_ids_from_log(CLOSING_LOGS_PATH))
    if CLOSING_LOGS_BY_DATE_DIR.exists():
        for log_path in CLOSING_LOGS_BY_DATE_DIR.glob("*.txt"):
            closed_ids.update(_load_order_ids_from_log(log_path))
    return closed_ids


def _extract_order_id(item: dict) -> str:
    return str(
        item.get("ord_ID") or
        item.get("ordid") or
        item.get("ord_id") or
        item.get("orderId") or
        item.get("order_id") or
        ""
    ).strip()


# ── Output writers ───────────────────────────────────────────────────────────

FIELDNAMES = [
    "order_id", "task_id", "source_status",
    "subject", "last_processed_at"
]


def save_csv(records: list):
    """Write / append records to data.csv."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    file_exists = CSV_PATH.exists()

    # Collect all keys across all records for dynamic columns
    all_keys = list(FIELDNAMES)
    for rec in records:
        for k in rec:
            if k not in all_keys:
                all_keys.append(k)

    with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        if not file_exists:
            writer.writeheader()
        writer.writerows(records)

    print(f"[GetTask] CSV saved/updated: {CSV_PATH}  (+{len(records)} rows)")


def save_excel(records: list):
    """Write all records from data.csv into data.xlsx (full overwrite each run)."""
    if not OPENPYXL_AVAILABLE:
        print("[GetTask] WARNING: openpyxl not installed – skipping Excel output. "
              "Run: pip install openpyxl")
        return

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Read current CSV (includes previously saved + new records)
    all_rows = []
    if CSV_PATH.exists():
        with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or FIELDNAMES
            all_rows = list(reader)
    else:
        fieldnames = FIELDNAMES

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GetTask Data"
    ws.append(list(fieldnames))
    for row in all_rows:
        ws.append([row.get(col, "") for col in fieldnames])

    wb.save(EXCEL_PATH)
    print(f"[GetTask] Excel saved: {EXCEL_PATH}  ({len(all_rows)} rows total)")


def save_payload(raw_payload):
    """Save the raw server payload to payload.json."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(PAYLOAD_PATH, "w", encoding="utf-8") as f:
        json.dump(raw_payload, f, indent=2, ensure_ascii=False)
    count = len(raw_payload) if isinstance(raw_payload, list) else 1
    print(f"[GetTask] Payload JSON saved: {PAYLOAD_PATH}  ({count} items)")


def save_query_txt(request_url: str, params: dict, headers: dict, actual_url: str):
    """Save query.txt – shows exactly what was sent to the server (credentials masked)."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    # Mask any header value that looks like a credential
    safe_headers = {k: ("[REDACTED]" if any(s in k.upper() for s in ("KEY", "AUTH", "TOKEN", "SECRET", "PASS")) else v) for k, v in headers.items()}
    with open(QUERY_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write(f"Generated at : {get_now_iso()}\n")
        f.write("\n")
        f.write("=== REQUEST URL (base) ===\n")
        f.write(f"{request_url}\n")
        f.write("\n")
        f.write("=== QUERY PARAMETERS SENT ===\n")
        for k, v in params.items():
            f.write(f"  {k} = {v}\n")
        f.write("\n")
        f.write("=== FULL URL (as prepared by requests) ===\n")
        f.write(f"{actual_url}\n")
        f.write("\n")
        f.write("=== HEADERS SENT (credentials masked) ===\n")
        for k, v in safe_headers.items():
            f.write(f"  {k}: {v}\n")
    print(f"[GetTask] Query saved : {QUERY_PATH}")


def save_one_record_to_csv(record: dict):
    """Append a single record to data.csv immediately."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    file_exists = CSV_PATH.exists()

    # Build column order: canonical first, then any extra
    all_keys = list(FIELDNAMES)
    for k in record:
        if k not in all_keys:
            all_keys.append(k)

    with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        if not file_exists:
            writer.writeheader()
        writer.writerow(record)


def save_empty_snapshot():
    """Create empty CSV/XLSX outputs so downstream steps do not read stale rows."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()

    if OPENPYXL_AVAILABLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GetTask Data"
        ws.append(list(FIELDNAMES))
        wb.save(EXCEL_PATH)


def rebuild_excel_from_csv():
    """Rebuild data.xlsx from the current data.csv (called after each save)."""
    if not OPENPYXL_AVAILABLE:
        return
    if not CSV_PATH.exists():
        return
    with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames or FIELDNAMES
        all_rows = list(reader)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GetTask Data"
    ws.append(list(fieldnames))
    for row in all_rows:
        ws.append([row.get(col, "") for col in fieldnames])
    wb.save(EXCEL_PATH)
    print(f"  📊 Excel updated: {len(all_rows)} rows total")


# ── Core fetch ───────────────────────────────────────────────────────────────

def fetch_all_tasks(request_url: str, params: dict, headers: dict, max_items: int = 0):
    """
    Fetch tasks from the server. Handles both paginated and non-paginated responses.
    Returns (all_items, actual_url_used, final_params_used).

    Strategy:
      1. Try a page-based loop (page=1,2,3…) until an empty page is returned.
      2. If the API doesn't support pagination, the first response is everything.
    """
    all_items      = []
    page           = 1
    max_pages      = 500
    page_size_hint = 100    # Tell server we want up to 100 per page if it supports it
    actual_url     = request_url   # will be updated after first real request
    final_params   = dict(params)
    previous_page_ids = None

    while True:
        if page > max_pages:
            print(f"[GetTask] Safety stop: reached max pages ({max_pages}).")
            break

        paged_params = dict(params)
        paged_params["page"]     = page
        paged_params["pageSize"] = page_size_hint
        paged_params["limit"]    = page_size_hint
        paged_params["offset"]   = (page - 1) * page_size_hint

        print(f"[GetTask] Fetching page {page} …")
        try:
            resp = requests.get(
                request_url,
                params=paged_params,
                headers=headers,
                timeout=TIMEOUT_SECONDS
            )
            resp.raise_for_status()
            # Capture the exact URL requests actually sent
            if page == 1:
                actual_url   = resp.url
                final_params = paged_params
            payload = resp.json()
        except Exception as e:
            print(f"[GetTask] Request failed on page {page}: {e}")
            break

        # Normalise to list
        if isinstance(payload, dict):
            items = (
                payload.get("data")
                or payload.get("results")
                or payload.get("tasks")
                or payload.get("items")
                or []
            )
            if not items and page == 1:
                items = [payload] if payload else []
        elif isinstance(payload, list):
            items = payload
        else:
            items = []

        if not items:
            break

        current_page_ids = tuple(_extract_order_id(item) for item in items if isinstance(item, dict))
        if previous_page_ids is not None and current_page_ids and current_page_ids == previous_page_ids:
            print("[GetTask] Safety stop: detected repeated page payload; pagination likely ignored by API.")
            break

        previous_page_ids = current_page_ids if current_page_ids else previous_page_ids

        all_items.extend(items)
        print(f"[GetTask] Page {page}: received {len(items)} items  (total so far: {len(all_items)})")

        if max_items > 0 and len(all_items) >= max_items:
            all_items = all_items[:max_items]
            print(f"[GetTask] Reached fetch cap ({max_items}) — stopping pagination.")
            break

        if len(items) < page_size_hint:
            break

        page += 1

    return all_items, actual_url, final_params


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    configure_console_encoding()

    parser = argparse.ArgumentParser(description="GetTask – fetch open CRM tasks")
    parser.add_argument("--force",  action="store_true",
                        help="Ignore logs.txt and reprocess all order IDs")
    parser.add_argument("--limit",  type=int, default=0,
                        help="Cap how many NEW records to save (0 = unlimited)")
    parser.add_argument("--fetch-limit", type=int, default=100,
                        help="Cap how many records to fetch from server (default: 100, 0 = unlimited)")
    args = parser.parse_args()

    load_dotenv_file()

    raw_api_url   = _required_env("API_URL_TASK_OPENED")
    task_subject  = _required_env("TASK_SUBJECT")
    api_key_header = _required_env("API_KEY_HEADER")
    api_key_value  = _required_env("API_KEY_VALUE")

    headers     = {api_key_header: api_key_value, "Accept": "application/json"}
    request_url = _normalized_url(raw_api_url).split("?")[0]
    params      = {"trsubject": task_subject}

    print(f"[GetTask] Fetching open tasks from: {request_url}")
    print(f"[GetTask] Subject filter: {task_subject}")

    # ── 1. Fetch ALL items from server ───────────────────────────────────────
    fetch_cap = args.fetch_limit if args.fetch_limit > 0 else 0
    raw_payload, actual_url, final_params = fetch_all_tasks(request_url, params, headers, max_items=fetch_cap)

    print(f"")
    print(f"┌─────────────────────────────────────────────────────────┐")
    print(f"│  SERVER RESPONSE SUMMARY                                │")
    print(f"│  Total items received : {len(raw_payload):<33}│")
    print(f"└─────────────────────────────────────────────────────────┘")

    # Save query.txt – exact request that was sent
    save_query_txt(request_url, final_params, headers, actual_url)

    # Always materialize output files so downstream stages can detect an empty run
    # instead of failing on missing files.
    if not raw_payload:
        save_empty_snapshot()

    if not raw_payload:
        print("[GetTask] No tasks returned – nothing to do.")
        return

    # Always save the raw payload exactly as received
    save_payload(raw_payload)

    # ── 2. Load already-processed IDs from logs.txt ──────────────────────────
    logged_ids = set() if args.force else load_logged_ids()
    closed_order_ids = set() if args.force else load_closed_order_ids()
    if logged_ids:
        print(f"[GetTask] Loaded {len(logged_ids)} already-processed IDs from logs.txt")
    else:
        if LOGS_PATH.exists() and not args.force:
            print("[GetTask] logs.txt exists but is empty – will process all items")
        elif not LOGS_PATH.exists():
            print("[GetTask] logs.txt not found – will process all items")

    # ── 3. Process each item (save CSV + Excel immediately after each) ────────
    if closed_order_ids:
        print(f"[GetTask] Excluding {len(closed_order_ids)} already-closed order IDs from ClosingTask logs")

    new_count     = 0
    skipped_count = 0
    total_items   = len(raw_payload)

    print(f"\n{'='*60}")
    print(f"  LIVE PROCESSING  –  {total_items} items from server")
    print(f"{'='*60}\n")

    for idx, item in enumerate(raw_payload, start=1):
        if not isinstance(item, dict):
            continue

        # Resolve order_id from known field names
        order_id = _normalize_order_id(_extract_order_id(item))

        task_id = str(item.get("trID") or item.get("task_id") or "").strip()

        print(f"[{idx}/{total_items}] ─────────────────────────────────────")
        print(f"  Order ID : {order_id or '(not found)'}")
        print(f"  Task  ID : {task_id or '(none)'}")

        if not order_id:
            print(f"  ⚠  WARNING: no order_id in response – skipping")
            print(f"     Raw item: {item}")
            continue

        if order_id in logged_ids and not args.force:
            print(f"  ⏭  SKIP – already in logs.txt")
            skipped_count += 1
            continue

        if order_id in closed_order_ids and not args.force:
            print("  SKIP - already closed by ClosingTask")
            skipped_count += 1
            continue

        # Show key server fields
        status  = item.get("trStatus", "Open")
        subject = item.get("trSubject", task_subject)
        print(f"  Status   : {status}")
        print(f"  Subject  : {subject}")

        # Show any extra interesting fields from server response
        extra_fields = {k: v for k, v in item.items()
                        if k not in ("ord_ID", "ordid", "ord_id", "orderId", "order_id",
                                     "trID", "task_id", "trStatus", "trSubject")}
        if extra_fields:
            print(f"  Extra fields from server:")
            for k, v in extra_fields.items():
                print(f"    {k}: {v}")

        record = {
            "order_id":          order_id,
            "task_id":           task_id,
            "source_status":     status,
            "subject":           subject,
            "last_processed_at": get_now_iso(),
        }
        for k, v in item.items():
            if k not in record:
                record[k] = v

        # ── IMMEDIATE save: CSV row + full Excel rebuild ──────────────────
        save_one_record_to_csv(record)
        rebuild_excel_from_csv()
        logged_ids.add(order_id)
        append_logged_id(order_id)
        new_count += 1
        print(f"  ✓ SAVED to CSV & Excel  (total saved so far: {new_count})")

        if args.limit > 0 and new_count >= args.limit:
            print(f"\n[GetTask] Reached --limit={args.limit}; stopping early.")
            break

    print(f"\n{'='*60}")
    print(f"  DONE  –  New saved: {new_count}  |  Skipped: {skipped_count}")
    print(f"{'='*60}")
    print(f"\n[GetTask] Output folder : {OUTPUT_DIR}")
    print(f"[GetTask] Files created :")
    for f in [CSV_PATH, EXCEL_PATH, PAYLOAD_PATH, LOGS_PATH, QUERY_PATH]:
        exists = "✓" if f.exists() else "✗"
        print(f"  {exists}  {f.name}")


if __name__ == "__main__":
    main()
